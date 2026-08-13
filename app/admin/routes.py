"""
Admin routes.

Every view here is protected by `@login_required`. Add new admin pages by
declaring more routes in this file (or by creating additional blueprints for
larger feature areas and registering them in `app/__init__.py`).
"""
import pyodbc
import secrets
from zipfile import BadZipFile
from datetime import date
from io import BytesIO

from flask import (
    current_app,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from werkzeug.security import generate_password_hash

from app.admin import admin_bp
from app.auth.repository import authenticate
from app.db import get_connection
from app.db_settings import apply_db_settings, save_db_settings, settings_have_password
from app.utils.decorators import login_required, role_required


ALLOWED_EXCEL_MIMETYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}
ALLOWED_IMPORT_ROLES = {"Admin", "User"}


@admin_bp.route("/login", methods=["GET", "POST"])
def login():
    """Admin sign-in page."""
    if session.get("logged_in"):
        if session.get("role") == "Admin":
            return redirect(url_for("admin.dashboard"))
        return redirect(url_for("user.dashboard"))

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")

        try:
            user = authenticate(username, password)
        except pyodbc.Error as exc:
            current_app.logger.error("Database error during admin login: %s", exc)
            flash(
                "Could not reach the database. Please contact the administrator.",
                "error",
            )
            return render_template("admin/login.html"), 503

        if user and user["role"] == "Admin":
            session.clear()
            session["logged_in"] = True
            session["user_id"] = user["id"]
            session["username"] = user["username"]
            session["role"] = user["role"]
            session.permanent = True
            return redirect(url_for("admin.dashboard"))

        flash("Admin access is required.", "error")
        return render_template("admin/login.html"), 401

    return render_template("admin/login.html")


@admin_bp.route("/" )
@admin_bp.route("/dashboard")
@login_required
@role_required("Admin")
def dashboard():
    """Admin overview page — passes stat counts for the KPI cards."""
    total_users = active_users = admin_users = total_employees = None
    try:
        with get_connection() as conn:
            cur = conn.cursor()
            cur.execute("SELECT COUNT(*), SUM(CASE WHEN IsActive=1 THEN 1 ELSE 0 END), SUM(CASE WHEN Role='Admin' THEN 1 ELSE 0 END) FROM dbo.Users")
            row = cur.fetchone()
            if row:
                total_users, active_users, admin_users = row
            # Employees table may not exist yet — handle gracefully
            try:
                cur.execute("SELECT COUNT(*) FROM dbo.Employees")
                emp_row = cur.fetchone()
                total_employees = emp_row[0] if emp_row else 0
            except Exception:
                total_employees = None
    except Exception:
        pass  # Stats are display-only; a DB error must not break the page

    return render_template(
        "admin/dashboard.html",
        username=session.get("username", "Admin"),
        total_users=total_users,
        active_users=active_users,
        admin_users=admin_users,
        total_employees=total_employees,
    )


@admin_bp.route("/reports")
@login_required
@role_required("Admin")
def reports():
    """Reports page with Excel import tools (Admin only)."""
    return render_template(
        "admin/reports.html",
        username=session.get("username", "Admin"),
    )


def _clean_cell(value):
    if value is None:
        return ""
    return str(value).strip()


def _normalize_header(value):
    return _clean_cell(value).lower()


def _normalize_role(value):
    role = _clean_cell(value).title()
    if role not in ALLOWED_IMPORT_ROLES:
        return ""
    return role


def _row_value(row, header_map, column_name):
    index = header_map.get(column_name)
    if index is None or index >= len(row):
        return ""
    return row[index]


def _placeholder_password_hash():
    return generate_password_hash(secrets.token_urlsafe(32))


@admin_bp.route("/reports/import", methods=["POST"])
@login_required
@role_required("Admin")
def import_report_users():
    """Import users from an uploaded Excel workbook (Admin only)."""
    upload = request.files.get("excel_file")
    if not upload or not upload.filename:
        flash("Please choose an Excel file to import.", "warning")
        return redirect(url_for("admin.reports"))

    filename = upload.filename.strip().lower()
    if not filename.endswith(".xlsx"):
        flash("Only .xlsx Excel files are supported.", "warning")
        return redirect(url_for("admin.reports"))

    if upload.mimetype not in ALLOWED_EXCEL_MIMETYPES:
        flash("The uploaded file does not look like a valid .xlsx file.", "warning")
        return redirect(url_for("admin.reports"))

    try:
        workbook = load_workbook(upload.stream, read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, OSError, ValueError):
        flash("The uploaded file could not be opened as a valid Excel workbook.", "error")
        return redirect(url_for("admin.reports"))

    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        flash("The Excel file is empty.", "warning")
        return redirect(url_for("admin.reports"))

    header_map = {
        _normalize_header(header): index
        for index, header in enumerate(headers)
        if _normalize_header(header)
    }
    required_columns = {"username", "role"}
    missing_columns = sorted(required_columns - set(header_map))
    if missing_columns:
        flash(
            "Missing required column(s): " + ", ".join(missing_columns),
            "warning",
        )
        return redirect(url_for("admin.reports"))

    imported_count = 0
    skipped = []
    seen_usernames = set()

    with get_connection() as conn:
        cur = conn.cursor()

        for excel_row_number, row in enumerate(rows, start=2):
            username = _clean_cell(_row_value(row, header_map, "username"))
            role = _normalize_role(_row_value(row, header_map, "role"))
            password = _clean_cell(_row_value(row, header_map, "password"))

            if not username:
                skipped.append(f"row {excel_row_number} missing Username")
                continue
            if len(username) > 100:
                skipped.append(f"row {excel_row_number} Username is too long")
                continue
            if not role:
                skipped.append(f"row {excel_row_number} has invalid Role")
                continue
            if password and len(password) < 6:
                skipped.append(f"row {excel_row_number} password is too short")
                continue

            username_key = username.lower()
            if username_key in seen_usernames:
                skipped.append(f"row {excel_row_number} duplicate Username")
                continue
            seen_usernames.add(username_key)

            cur.execute("SELECT Id FROM Users WHERE Username = ?", username)
            existing = cur.fetchone()

            if existing:
                if password:
                    cur.execute(
                        "UPDATE Users SET PasswordHash = ?, Role = ?, IsActive = 1 "
                        "WHERE Username = ?",
                        generate_password_hash(password), role, username,
                    )
                else:
                    cur.execute(
                        "UPDATE Users SET Role = ? WHERE Username = ?",
                        role, username,
                    )
            else:
                if password:
                    password_hash = generate_password_hash(password)
                    is_active = 1
                else:
                    password_hash = _placeholder_password_hash()
                    is_active = 0

                cur.execute(
                    "INSERT INTO Users (Username, PasswordHash, Role, IsActive) "
                    "VALUES (?, ?, ?, ?)",
                    username, password_hash, role, is_active,
                )

            imported_count += 1

        conn.commit()

    if skipped:
        preview = "; ".join(skipped[:5])
        if len(skipped) > 5:
            preview += f"; and {len(skipped) - 5} more"
        flash(
            f"{imported_count} rows imported, {len(skipped)} skipped: {preview}.",
            "warning",
        )
    else:
        flash(f"{imported_count} rows imported successfully.", "success")

    return redirect(url_for("admin.reports"))


@admin_bp.route("/users")
@login_required
@role_required("Admin")
def users_page():
    """List all users (Admin only).

    Joins to Employees to surface the linked department and show an
    'Employee account' badge vs 'Staff account' for manually-created rows.
    """
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(
            """
            SELECT
                u.Id,
                u.Username,
                u.Role,
                u.IsActive,
                u.EmployeeId,
                e.Name        AS EmployeeName,
                e.Department  AS EmployeeDepartment
            FROM dbo.Users u
            LEFT JOIN dbo.Employees e ON e.Id = u.EmployeeId
            ORDER BY u.Id
            """
        )
        rows = cur.fetchall()

    users = [
        {
            "id":                  r[0],
            "username":            r[1],
            "role":                r[2],
            "is_active":           bool(r[3]),
            "employee_id":         r[4],
            "employee_name":       r[5],
            "employee_department": r[6],
        }
        for r in rows
    ]

    return render_template(
        "admin/users.html",
        username=session.get("username", "Admin"),
        users=users,
    )


@admin_bp.route("/users/export")
@login_required
@role_required("Admin")
def export_users():
    """Export users to an Excel workbook (Admin only)."""
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(
            "SELECT Id, Username, Role, IsActive, CreatedAt FROM Users ORDER BY Id"
        )
        rows = cur.fetchall()

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Users"
    worksheet.append(["Id", "Username", "Role", "IsActive", "CreatedAt"])

    for row in rows:
        worksheet.append(
            [
                row[0],
                row[1],
                row[2],
                "Yes" if row[3] else "No",
                row[4],
            ]
        )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"users_export_{date.today().isoformat()}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _count_active_admins(exclude_user_id=None):
    with get_connection() as conn:
        cur = conn.cursor()
        if exclude_user_id is None:
            cur.execute(
                "SELECT COUNT(1) FROM Users WHERE Role = 'Admin' AND IsActive = 1"
            )
            return int(cur.fetchone()[0])
        cur.execute(
            "SELECT COUNT(1) FROM Users WHERE Role = 'Admin' AND IsActive = 1 AND Id <> ?",
            exclude_user_id,
        )
        return int(cur.fetchone()[0])


@admin_bp.route("/users/<int:user_id>/toggle-role", methods=["POST"])
@login_required
@role_required("Admin")
def toggle_user_role(user_id: int):
    """Promote to Admin or demote to User (Admin only).

    Safety: prevent demoting the last active admin.
    """
    action = request.form.get("action", "").strip().lower()
    if action not in {"promote", "demote"}:
        flash("Invalid role action.", "error")
        return redirect(f"{admin_bp.url_prefix}/users")

    with get_connection() as conn:
        cur = conn.cursor()

        # If demoting/promoting, we update Role and ensure IsActive stays 1.
        if action == "demote":
            # Safety: if this user is the last active admin, block.
            # Exclude current user id from count.
            remaining = _count_active_admins(exclude_user_id=user_id)
            if remaining <= 0:
                flash("Cannot demote: you must keep at least one active administrator.", "warning")
                return redirect(f"{admin_bp.url_prefix}/users")
            cur.execute(
                "UPDATE Users SET Role = 'User' WHERE Id = ?",
                user_id,
            )
        else:
            cur.execute(
                "UPDATE Users SET Role = 'Admin' WHERE Id = ?",
                user_id,
            )

        conn.commit()

    flash("User role updated.", "success")
    return redirect(f"{admin_bp.url_prefix}/users")


@admin_bp.route("/users/<int:user_id>/disable", methods=["POST"])
@login_required
@role_required("Admin")
def disable_user(user_id: int):
    """Disable (IsActive=0) a user (Admin only).

    Safety: prevent disabling the last active admin.
    """
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(
            "SELECT Role FROM Users WHERE Id = ?",
            user_id,
        )
        row = cur.fetchone()
        if not row:
            flash("User not found.", "error")
            return redirect(f"{admin_bp.url_prefix}/users")

        role = row[0]

        if role == "Admin":
            remaining = _count_active_admins(exclude_user_id=user_id)
            if remaining <= 0:
                flash("Cannot disable: you must keep at least one active administrator.", "warning")
                return redirect(f"{admin_bp.url_prefix}/users")

        cur.execute(
            "UPDATE Users SET IsActive = 0 WHERE Id = ?",
            user_id,
        )
        conn.commit()

    flash("User disabled.", "success")
    return redirect(f"{admin_bp.url_prefix}/users")


# ---------------------------------------------------------------------------
# Settings — database connection configuration
# ---------------------------------------------------------------------------

@admin_bp.route("/settings", methods=["GET"])
@login_required
@role_required("Admin")
def settings():
    """Show the DB settings form, pre-filled from current app config."""
    cfg = current_app.config
    return render_template(
        "admin/settings.html",
        username=session.get("username", "Admin"),
        # Form field values — password is never sent back to the browser.
        db_server=cfg.get("DB_SERVER", ""),
        db_name=cfg.get("DB_NAME", ""),
        db_driver=cfg.get("DB_DRIVER", ""),
        db_username=cfg.get("DB_USERNAME", ""),
        db_trusted=cfg.get("DB_TRUSTED_CONNECTION", True),
        db_encrypt=cfg.get("DB_ENCRYPT", "yes"),
        db_trust_cert=cfg.get("DB_TRUST_SERVER_CERTIFICATE", "yes"),
        db_timeout=cfg.get("DB_TIMEOUT", 5),
        has_password=settings_have_password(),
    )


@admin_bp.route("/settings", methods=["POST"])
@login_required
@role_required("Admin")
def settings_save():
    """Save DB settings to db_settings.json and hot-reload them into config."""
    trusted = bool(request.form.get("db_trusted_connection"))

    server = request.form.get("db_server", "").strip()
    db_name = request.form.get("db_name", "").strip()
    driver = request.form.get("db_driver", "").strip()
    username = request.form.get("db_username", "").strip()
    password = request.form.get("db_password", "")  # intentionally no strip
    encrypt = request.form.get("db_encrypt", "yes").strip()
    trust_cert = request.form.get("db_trust_server_certificate", "yes").strip()
    timeout_raw = request.form.get("db_timeout", "5").strip()

    # --- Basic validation -------------------------------------------------
    errors = []
    if not server:
        errors.append("DB Server is required.")
    if not db_name:
        errors.append("Database Name is required.")
    if not driver:
        errors.append("ODBC Driver is required.")
    if not trusted and not username:
        errors.append("Username is required when not using Windows Authentication.")
    try:
        timeout = int(timeout_raw)
        if timeout < 1:
            raise ValueError
    except ValueError:
        errors.append("Timeout must be a positive integer (seconds).")
        timeout = 5

    if errors:
        for msg in errors:
            flash(msg, "error")
        return redirect(url_for("admin.settings"))

    # --- Build the dict to persist ----------------------------------------
    to_save = {
        "DB_SERVER": server,
        "DB_NAME": db_name,
        "DB_DRIVER": driver,
        "DB_USERNAME": username,
        "DB_TRUSTED_CONNECTION": trusted,
        "DB_ENCRYPT": encrypt,
        "DB_TRUST_SERVER_CERTIFICATE": trust_cert,
        "DB_TIMEOUT": timeout,
    }

    # Only update the stored password when the user actually typed one.
    # An empty field means "keep whatever is already saved".
    if password:
        to_save["DB_PASSWORD"] = password

    # --- Persist + hot-reload into the live Flask config ------------------
    save_db_settings(to_save)
    apply_db_settings(current_app._get_current_object())

    flash("Settings saved successfully.", "success")
    return redirect(url_for("admin.settings"))


@admin_bp.route("/settings/test-connection", methods=["POST"])
@login_required
@role_required("Admin")
def settings_test_connection():
    """Try opening a pyodbc connection with the current config and report the result."""
    try:
        with get_connection() as conn:
            cur = conn.cursor()
            cur.execute("SELECT 1")          # cheapest possible round-trip
            cur.fetchone()
        flash(
            f"Connection successful — reached "
            f"[{current_app.config.get('DB_SERVER')}] / "
            f"{current_app.config.get('DB_NAME')}.",
            "success",
        )
    except pyodbc.Error as exc:
        # Surface the exact ODBC error so the admin can diagnose it.
        flash(f"Connection failed: {exc}", "error")
    except Exception as exc:
        flash(f"Unexpected error: {exc}", "error")

    return redirect(url_for("admin.settings"))


# ---------------------------------------------------------------------------
# Employees import — separate from the Users import above
# ---------------------------------------------------------------------------

def _parse_salary(raw: str) -> int | None:
    """Convert a raw cell value to a non-negative integer salary.

    Returns None if the value is missing, non-numeric, or negative.
    """
    cleaned = raw.strip().replace(",", "").replace(" ", "")
    if not cleaned:
        return None
    try:
        value = int(float(cleaned))   # accept "50000.0" from Excel number cells
        if value < 0:
            return None
        return value
    except (ValueError, OverflowError):
        return None


@admin_bp.route("/reports/import-employees", methods=["POST"])
@login_required
@role_required("Admin")
def import_employees():
    """Import employee CRM records from an uploaded Excel workbook (Admin only).

    Required columns (case-insensitive): Name, Department, Salary.
    Each valid row is inserted into dbo.Employees.
    This route is completely independent of import_report_users.
    """
    upload = request.files.get("employees_file")
    if not upload or not upload.filename:
        flash("Please choose an Excel file to import.", "warning")
        return redirect(url_for("admin.reports"))

    if not upload.filename.strip().lower().endswith(".xlsx"):
        flash("Only .xlsx Excel files are supported.", "warning")
        return redirect(url_for("admin.reports"))

    if upload.mimetype not in ALLOWED_EXCEL_MIMETYPES:
        flash("The uploaded file does not look like a valid .xlsx file.", "warning")
        return redirect(url_for("admin.reports"))

    try:
        workbook = load_workbook(upload.stream, read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, OSError, ValueError):
        flash(
            "The uploaded file could not be opened as a valid Excel workbook.",
            "error",
        )
        return redirect(url_for("admin.reports"))

    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)

    headers = next(rows, None)
    if not headers:
        flash("The Excel file is empty.", "warning")
        return redirect(url_for("admin.reports"))

    header_map = {
        _normalize_header(h): idx
        for idx, h in enumerate(headers)
        if _normalize_header(h)
    }

    required_columns = {"name", "department", "salary"}
    missing_columns = sorted(required_columns - set(header_map))
    if missing_columns:
        flash(
            "Missing required column(s): " + ", ".join(c.title() for c in missing_columns),
            "warning",
        )
        return redirect(url_for("admin.reports"))

    imported_count = 0
    skipped = []

    with get_connection() as conn:
        cur = conn.cursor()

        for excel_row_number, row in enumerate(rows, start=2):
            name = _clean_cell(_row_value(row, header_map, "name"))
            department = _clean_cell(_row_value(row, header_map, "department"))
            salary_raw = _clean_cell(_row_value(row, header_map, "salary"))

            # --- Row-level validation -----------------------------------
            if not name:
                skipped.append(f"row {excel_row_number} missing Name")
                continue
            if len(name) > 100:
                skipped.append(f"row {excel_row_number} Name is too long (max 100)")
                continue
            if not department:
                skipped.append(f"row {excel_row_number} missing Department")
                continue
            if len(department) > 100:
                skipped.append(
                    f"row {excel_row_number} Department is too long (max 100)"
                )
                continue

            salary = _parse_salary(salary_raw)
            if salary is None:
                skipped.append(
                    f"row {excel_row_number} Salary is missing or not a valid "
                    f"non-negative integer"
                )
                continue

            # --- Insert -------------------------------------------------
            cur.execute(
                "INSERT INTO dbo.Employees (Name, Department, Salary) "
                "VALUES (?, ?, ?)",
                name, department, salary,
            )
            imported_count += 1

        conn.commit()

    if skipped:
        preview = "; ".join(skipped[:5])
        if len(skipped) > 5:
            preview += f"; and {len(skipped) - 5} more"
        flash(
            f"{imported_count} employee(s) imported, "
            f"{len(skipped)} skipped: {preview}.",
            "warning",
        )
    else:
        flash(f"{imported_count} employee(s) imported successfully.", "success")

    return redirect(url_for("admin.reports"))


# ---------------------------------------------------------------------------
# Employees CRUD + Export
# ---------------------------------------------------------------------------

import re as _re   # used by _generate_username only; aliased to avoid shadowing


def _generate_username(name: str, cur) -> str:
    """Derive a unique username from an employee name.

    Algorithm:
      1. Lowercase the name, keep only ASCII letters (strips accents, spaces,
         punctuation) → base slug, e.g. "Karim Mansour" → "karimmansour".
         If that would be empty, fall back to "user".
      2. Find all existing usernames that start with the slug in one query.
      3. Return the slug if unclaimed, else slug+2, slug+3, ... (first gap).

    The cursor must be an open pyodbc cursor on the same connection.
    """
    slug = _re.sub(r"[^a-z]", "", name.lower()) or "user"
    # Truncate so slug + up to 3 digits fits within the 100-char Username limit
    slug = slug[:96]

    cur.execute(
        "SELECT Username FROM dbo.Users WHERE Username LIKE ?",
        slug + "%",
    )
    taken = {row[0].lower() for row in cur.fetchall()}

    if slug not in taken:
        return slug

    counter = 2
    while True:
        candidate = f"{slug}{counter}"
        if candidate not in taken:
            return candidate
        counter += 1


def _create_linked_user(employee_id: int, employee_name: str, cur) -> str:
    """Insert a new Users row linked to the given Employee.

    Returns the plain-text temporary password (shown once in the flash
    message — never stored or logged anywhere after this function returns).

    The cursor must be an open, writable pyodbc cursor.
    The caller is responsible for calling conn.commit().
    """
    username = _generate_username(employee_name, cur)
    temp_password = secrets.token_urlsafe(10)   # 10-char URL-safe random string
    password_hash = generate_password_hash(temp_password)

    cur.execute(
        "INSERT INTO dbo.Users "
        "  (Username, PasswordHash, Role, IsActive, EmployeeId) "
        "VALUES (?, ?, 'User', 1, ?)",
        username, password_hash, employee_id,
    )
    return username, temp_password

def _get_employee_or_404(employee_id: int) -> dict:
    """Fetch a single employee row by Id, or raise 404 if not found."""
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(
            "SELECT Id, Name, Department, Salary, CreatedAt "
            "FROM dbo.Employees WHERE Id = ?",
            employee_id,
        )
        row = cur.fetchone()
    if not row:
        from flask import abort
        abort(404)
    return {
        "id": row[0],
        "name": row[1],
        "department": row[2],
        "salary": row[3],
        "created_at": row[4],
    }


def _validate_employee_form(name: str, department: str, salary_raw: str):
    """Validate the Create/Edit employee form fields.

    Returns (salary_int, errors) where errors is a list of strings.
    salary_int is None when validation fails.
    """
    errors = []

    name = name.strip()
    department = department.strip()

    if not name:
        errors.append("Name is required.")
    elif len(name) > 100:
        errors.append("Name must be 100 characters or fewer.")

    if not department:
        errors.append("Department is required.")
    elif len(department) > 100:
        errors.append("Department must be 100 characters or fewer.")

    salary = _parse_salary(salary_raw)
    if salary is None:
        errors.append("Salary must be a valid positive whole number.")
    elif salary == 0:
        errors.append("Salary must be greater than zero.")

    return salary, errors


# ---- List ----------------------------------------------------------------

@admin_bp.route("/employees")
@login_required
@role_required("Admin")
def employees_list():
    """List all employees (Admin only)."""
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(
            "SELECT Id, Name, Department, Salary, CreatedAt "
            "FROM dbo.Employees ORDER BY Id"
        )
        rows = cur.fetchall()

    employees = [
        {
            "id": r[0],
            "name": r[1],
            "department": r[2],
            "salary": r[3],
            "created_at": r[4],
        }
        for r in rows
    ]

    return render_template(
        "admin/employees.html",
        username=session.get("username", "Admin"),
        employees=employees,
    )


# ---- Create --------------------------------------------------------------

@admin_bp.route("/employees/new", methods=["GET"])
@login_required
@role_required("Admin")
def employee_new():
    """Show the Create Employee form."""
    return render_template(
        "admin/employee_form.html",
        username=session.get("username", "Admin"),
        form_title="New employee",
        action=url_for("admin.employee_create"),
        employee=None,
        errors=[],
    )


@admin_bp.route("/employees/new", methods=["POST"])
@login_required
@role_required("Admin")
def employee_create():
    """Handle Create Employee form submission.

    After inserting the Employee row, automatically creates a linked Users row
    with a generated username and a one-time temporary password that is shown
    in the flash message and never stored in plain text.
    """
    name = request.form.get("name", "").strip()
    department = request.form.get("department", "").strip()
    salary_raw = request.form.get("salary", "").strip()

    salary, errors = _validate_employee_form(name, department, salary_raw)

    if errors:
        for msg in errors:
            flash(msg, "error")
        return render_template(
            "admin/employee_form.html",
            username=session.get("username", "Admin"),
            form_title="New employee",
            action=url_for("admin.employee_create"),
            employee={"name": name, "department": department, "salary": salary_raw},
            errors=errors,
        )

    with get_connection() as conn:
        cur = conn.cursor()

        # 1. Insert the Employee row and retrieve its new Id.
        cur.execute(
            "INSERT INTO dbo.Employees (Name, Department, Salary) "
            "VALUES (?, ?, ?); SELECT SCOPE_IDENTITY();",
            name, department, salary,
        )
        cur.nextset()                          # advance to the SELECT result
        employee_id = int(cur.fetchone()[0])

        # 2. Create the linked User account (same connection, same transaction).
        linked_username, temp_password = _create_linked_user(
            employee_id, name, cur
        )

        conn.commit()

    # Show the temporary password ONCE — it is never stored in plain text
    # after this point and will not appear again.
    flash(
        f"Employee '{name}' created. "
        f"Login: username={linked_username}, "
        f"password={temp_password} — share this with them, it won't be shown again.",
        "success",
    )
    return redirect(url_for("admin.employees_list"))


# ---- Update --------------------------------------------------------------

@admin_bp.route("/employees/<int:employee_id>/edit", methods=["GET"])
@login_required
@role_required("Admin")
def employee_edit(employee_id: int):
    """Show the Edit Employee form pre-filled with current values."""
    employee = _get_employee_or_404(employee_id)
    return render_template(
        "admin/employee_form.html",
        username=session.get("username", "Admin"),
        form_title=f"Edit — {employee['name']}",
        action=url_for("admin.employee_update", employee_id=employee_id),
        employee=employee,
        errors=[],
    )


@admin_bp.route("/employees/<int:employee_id>/edit", methods=["POST"])
@login_required
@role_required("Admin")
def employee_update(employee_id: int):
    """Handle Edit Employee form submission.

    If the employee's Name has changed, regenerates the linked Users.Username
    to match (checking for duplicates).  Password is never touched.
    """
    existing = _get_employee_or_404(employee_id)

    name = request.form.get("name", "").strip()
    department = request.form.get("department", "").strip()
    salary_raw = request.form.get("salary", "").strip()

    salary, errors = _validate_employee_form(name, department, salary_raw)

    if errors:
        for msg in errors:
            flash(msg, "error")
        return render_template(
            "admin/employee_form.html",
            username=session.get("username", "Admin"),
            form_title="Edit employee",
            action=url_for("admin.employee_update", employee_id=employee_id),
            employee={
                "id": employee_id,
                "name": name,
                "department": department,
                "salary": salary_raw,
            },
            errors=errors,
        )

    with get_connection() as conn:
        cur = conn.cursor()

        # Update the Employee row itself.
        cur.execute(
            "UPDATE dbo.Employees SET Name = ?, Department = ?, Salary = ? "
            "WHERE Id = ?",
            name, department, salary, employee_id,
        )

        # If the Name changed, regenerate the linked user's Username.
        if name != existing["name"]:
            cur.execute(
                "SELECT Id, Username FROM dbo.Users WHERE EmployeeId = ?",
                employee_id,
            )
            linked = cur.fetchone()
            if linked:
                linked_user_id, old_username = linked[0], linked[1]
                new_username = _generate_username(name, cur)
                cur.execute(
                    "UPDATE dbo.Users SET Username = ? WHERE Id = ?",
                    new_username, linked_user_id,
                )

        conn.commit()

    flash(f"Employee '{name}' updated successfully.", "success")
    return redirect(url_for("admin.employees_list"))


# ---- Delete --------------------------------------------------------------

@admin_bp.route("/employees/<int:employee_id>/delete", methods=["POST"])
@login_required
@role_required("Admin")
def employee_delete(employee_id: int):
    """Delete an Employee row and deactivate its linked Users account.

    The Users row is set to IsActive=0 (soft-delete) rather than hard-deleted
    so the audit trail is preserved.  The EmployeeId FK value is left in place
    so the link is still visible in the Users table.
    """
    employee = _get_employee_or_404(employee_id)
    name = employee["name"]

    with get_connection() as conn:
        cur = conn.cursor()

        # Deactivate the linked user account first (preserves audit trail).
        cur.execute(
            "UPDATE dbo.Users SET IsActive = 0 "
            "WHERE EmployeeId = ?",
            employee_id,
        )

        # Now delete the Employee row.
        # The FK allows NULL so the Users row stays after this delete.
        cur.execute(
            "UPDATE dbo.Users SET EmployeeId = NULL WHERE EmployeeId = ?",
            employee_id,
        )
        cur.execute("DELETE FROM dbo.Employees WHERE Id = ?", employee_id)
        conn.commit()

    flash(
        f"Employee '{name}' deleted. "
        f"The linked login account has been deactivated.",
        "success",
    )
    return redirect(url_for("admin.employees_list"))


# ---- Export --------------------------------------------------------------

@admin_bp.route("/employees/export")
@login_required
@role_required("Admin")
def export_employees():
    """Export all employees to an Excel workbook (Admin only)."""
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(
            "SELECT Id, Name, Department, Salary, CreatedAt "
            "FROM dbo.Employees ORDER BY Id"
        )
        rows = cur.fetchall()

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Employees"
    worksheet.append(["Id", "Name", "Department", "Salary", "CreatedAt"])

    for row in rows:
        worksheet.append([row[0], row[1], row[2], row[3], row[4]])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"employees_export_{date.today().isoformat()}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ---------------------------------------------------------------------------
# Users CRUD  (Create / Update / Delete)
# toggle_user_role and disable_user above are kept unchanged.
# ---------------------------------------------------------------------------

def _get_user_or_404(user_id: int) -> dict:
    """Fetch a single user row by Id, abort 404 if missing."""
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(
            "SELECT Id, Username, Role, IsActive FROM dbo.Users WHERE Id = ?",
            user_id,
        )
        row = cur.fetchone()
    if not row:
        from flask import abort
        abort(404)
    return {
        "id":        row[0],
        "username":  row[1],
        "role":      row[2],
        "is_active": bool(row[3]),
    }


def _validate_user_form(username: str, password: str, role: str,
                        is_edit: bool = False) -> list[str]:
    """Validate the Create / Edit user form.

    Returns a (possibly empty) list of error strings.
    On edit, password is optional — an empty string means "keep existing".
    """
    errors = []

    username = username.strip()
    role = role.strip()

    if not username:
        errors.append("Username is required.")
    elif len(username) > 100:
        errors.append("Username must be 100 characters or fewer.")

    if not is_edit and not password:
        errors.append("Password is required.")
    if password and len(password) < 6:
        errors.append("Password must be at least 6 characters.")

    if role not in ALLOWED_IMPORT_ROLES:
        errors.append("Role must be Admin or User.")

    return errors


# ---- Create ---------------------------------------------------------------

@admin_bp.route("/users/new", methods=["GET"])
@login_required
@role_required("Admin")
def user_new():
    """Show the Create User form."""
    return render_template(
        "admin/user_form.html",
        username=session.get("username", "Admin"),
        form_title="New user",
        action=url_for("admin.user_create"),
        user=None,
        errors=[],
    )


@admin_bp.route("/users/new", methods=["POST"])
@login_required
@role_required("Admin")
def user_create():
    """Handle Create User form submission."""
    new_username = request.form.get("username", "").strip()
    password     = request.form.get("password", "")
    role         = request.form.get("role", "User").strip()

    errors = _validate_user_form(new_username, password, role, is_edit=False)

    if errors:
        for msg in errors:
            flash(msg, "error")
        return render_template(
            "admin/user_form.html",
            username=session.get("username", "Admin"),
            form_title="New user",
            action=url_for("admin.user_create"),
            user={"username": new_username, "role": role},
            errors=errors,
        )

    # create_user() in repository handles hashing + duplicate check.
    from app.auth.repository import create_user
    result = create_user(new_username, password, role)

    if result is None:
        flash(f"Username '{new_username}' is already taken.", "error")
        return render_template(
            "admin/user_form.html",
            username=session.get("username", "Admin"),
            form_title="New user",
            action=url_for("admin.user_create"),
            user={"username": new_username, "role": role},
            errors=[],
        )

    flash(f"User '{new_username}' created successfully.", "success")
    return redirect(url_for("admin.users_page"))


# ---- Update ---------------------------------------------------------------

@admin_bp.route("/users/<int:user_id>/edit", methods=["GET"])
@login_required
@role_required("Admin")
def user_edit(user_id: int):
    """Show the Edit User form pre-filled with current values."""
    user = _get_user_or_404(user_id)
    return render_template(
        "admin/user_form.html",
        username=session.get("username", "Admin"),
        form_title=f"Edit — {user['username']}",
        action=url_for("admin.user_update", user_id=user_id),
        user=user,
        errors=[],
    )


@admin_bp.route("/users/<int:user_id>/edit", methods=["POST"])
@login_required
@role_required("Admin")
def user_update(user_id: int):
    """Handle Edit User form submission."""
    existing = _get_user_or_404(user_id)

    new_username = request.form.get("username", "").strip()
    password     = request.form.get("password", "")   # blank = keep existing
    role         = request.form.get("role", "User").strip()

    errors = _validate_user_form(new_username, password, role, is_edit=True)

    # Safety: if changing from Admin → User, ensure another active admin exists.
    if not errors and existing["role"] == "Admin" and role == "User":
        remaining = _count_active_admins(exclude_user_id=user_id)
        if remaining <= 0:
            errors.append(
                "Cannot change role: at least one active Admin must remain."
            )

    if errors:
        for msg in errors:
            flash(msg, "error")
        return render_template(
            "admin/user_form.html",
            username=session.get("username", "Admin"),
            form_title=f"Edit user",
            action=url_for("admin.user_update", user_id=user_id),
            user={"id": user_id, "username": new_username, "role": role},
            errors=errors,
        )

    # Check for username collision with a *different* user.
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(
            "SELECT Id FROM dbo.Users WHERE Username = ? AND Id <> ?",
            new_username, user_id,
        )
        if cur.fetchone():
            flash(f"Username '{new_username}' is already taken.", "error")
            return render_template(
                "admin/user_form.html",
                username=session.get("username", "Admin"),
                form_title=f"Edit user",
                action=url_for("admin.user_update", user_id=user_id),
                user={"id": user_id, "username": new_username, "role": role},
                errors=[],
            )

        if password:
            cur.execute(
                "UPDATE dbo.Users SET Username = ?, Role = ?, PasswordHash = ? "
                "WHERE Id = ?",
                new_username, role, generate_password_hash(password), user_id,
            )
        else:
            cur.execute(
                "UPDATE dbo.Users SET Username = ?, Role = ? WHERE Id = ?",
                new_username, role, user_id,
            )
        conn.commit()

    flash(f"User '{new_username}' updated successfully.", "success")
    return redirect(url_for("admin.users_page"))


# ---- Delete ---------------------------------------------------------------

@admin_bp.route("/users/<int:user_id>/delete", methods=["POST"])
@login_required
@role_required("Admin")
def user_delete(user_id: int):
    """Permanently delete a user row.

    Safety: blocks deletion of the last active Admin — same protection as
    disable_user / toggle_user_role.
    """
    user = _get_user_or_404(user_id)

    if user["role"] == "Admin" and user["is_active"]:
        remaining = _count_active_admins(exclude_user_id=user_id)
        if remaining <= 0:
            flash(
                "Cannot delete: at least one active Admin account must remain.",
                "warning",
            )
            return redirect(url_for("admin.users_page"))

    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute("DELETE FROM dbo.Users WHERE Id = ?", user_id)
        conn.commit()

    flash(f"User '{user['username']}' deleted.", "success")
    return redirect(url_for("admin.users_page"))
